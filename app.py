"""
RALCO Field Walkthrough — Backend v3
Sites → Contracts → Years → Assets hierarchy
"""

import io
import os
import re
import json
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

TEMPLATE_PATH     = os.path.join(os.path.dirname(__file__), "template.xlsx")
EPM_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "epm_template.xlsx")

# Fix DATABASE_URL for psycopg2 compatibility
raw_url = os.environ.get("DATABASE_URL", "")
if raw_url.startswith("postgres://"):
    raw_url = raw_url.replace("postgres://", "postgresql://", 1)
DATABASE_URL = raw_url or None

# ── DB helpers ────────────────────────────────────────────────
def get_db():
    import psycopg2
    return psycopg2.connect(DATABASE_URL, sslmode='require')

def init_db():
    if not DATABASE_URL:
        print("WARNING: No DATABASE_URL — cloud storage disabled")
        return
    try:
        conn = get_db(); cur = conn.cursor()

        # Sites table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sites (
                id          TEXT PRIMARY KEY,
                customer    TEXT,
                address     TEXT,
                contact     TEXT,
                phone       TEXT,
                notes       TEXT,
                device_id   TEXT,
                device_name TEXT,
                deleted     BOOLEAN DEFAULT FALSE,
                deleted_at  TIMESTAMP,
                created_at  TIMESTAMP DEFAULT NOW(),
                updated_at  TIMESTAMP DEFAULT NOW()
            )
        """)

        # Contracts table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS contracts (
                id              TEXT PRIMARY KEY,
                site_id         TEXT REFERENCES sites(id),
                contract_number INTEGER,
                tier            TEXT DEFAULT 'basic',
                status          TEXT DEFAULT 'Active',
                start_date      TEXT,
                end_date        TEXT,
                notes           TEXT,
                buildings       JSONB DEFAULT '[]',
                rooms           JSONB DEFAULT '[]',
                deleted         BOOLEAN DEFAULT FALSE,
                deleted_at      TIMESTAMP,
                created_at      TIMESTAMP DEFAULT NOW(),
                updated_at      TIMESTAMP DEFAULT NOW()
            )
        """)

        # Years table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS years (
                id              TEXT PRIMARY KEY,
                contract_id     TEXT REFERENCES contracts(id),
                site_id         TEXT REFERENCES sites(id),
                year_number     INTEGER,
                label           TEXT,
                work_date       TEXT,
                rep             TEXT,
                status          TEXT DEFAULT 'Scheduled',
                scoped_asset_ids JSONB DEFAULT '[]',
                deleted         BOOLEAN DEFAULT FALSE,
                deleted_at      TIMESTAMP,
                created_at      TIMESTAMP DEFAULT NOW(),
                updated_at      TIMESTAMP DEFAULT NOW()
            )
        """)

        # Assets table (site-level master)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS assets (
                id              TEXT PRIMARY KEY,
                site_id         TEXT REFERENCES sites(id),
                tag             TEXT,
                name            TEXT,
                location        TEXT,
                eq_type         TEXT,
                manufacturer    TEXT,
                model           TEXT,
                serial          TEXT,
                voltage         TEXT,
                criticality     TEXT,
                maint_status    TEXT,
                notes           TEXT,
                port_key        TEXT,
                service_history JSONB DEFAULT '[]',
                deleted         BOOLEAN DEFAULT FALSE,
                created_at      TIMESTAMP DEFAULT NOW(),
                updated_at      TIMESTAMP DEFAULT NOW()
            )
        """)

        # Meta table (tag counter etc)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS meta (
                key   TEXT PRIMARY KEY,
                value BIGINT
            )
        """)

        # Migrate old jobs table if it exists
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'jobs'
            )
        """)
        jobs_exists = cur.fetchone()[0]

        if jobs_exists:
            cur.execute("SELECT id, customer, data FROM jobs WHERE id NOT IN (SELECT id FROM sites)")
            old_jobs = cur.fetchall()
            for job_id, customer, data in old_jobs:
                if not data: continue
                # Create site from old job
                cur.execute("""
                    INSERT INTO sites (id, customer, address, notes, device_id, device_name, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                    ON CONFLICT (id) DO NOTHING
                """, (
                    job_id,
                    data.get('customer',''),
                    data.get('address',''),
                    data.get('notes',''),
                    data.get('deviceId',''),
                    data.get('deviceName',''),
                ))
                # Create contract from old job
                contract_id = 'c1_' + job_id
                cur.execute("""
                    INSERT INTO contracts (id, site_id, contract_number, tier, buildings, rooms, created_at, updated_at)
                    VALUES (%s, %s, 1, %s, %s, %s, NOW(), NOW())
                    ON CONFLICT (id) DO NOTHING
                """, (
                    contract_id,
                    job_id,
                    data.get('tier','basic'),
                    json.dumps(data.get('buildings',[])),
                    json.dumps(data.get('rooms',[])),
                ))
                # Create Year 1
                year_id = 'y1_' + job_id
                cur.execute("""
                    INSERT INTO years (id, contract_id, site_id, year_number, label, work_date, rep, status, created_at, updated_at)
                    VALUES (%s, %s, %s, 1, 'C1-Y1', %s, %s, 'In Progress', NOW(), NOW())
                    ON CONFLICT (id) DO NOTHING
                """, (
                    year_id,
                    contract_id,
                    job_id,
                    data.get('date',''),
                    data.get('rep',''),
                ))
                # Migrate assets
                for a in data.get('assets', []):
                    cur.execute("""
                        INSERT INTO assets (id, site_id, tag, name, location, eq_type, manufacturer,
                            model, serial, voltage, criticality, maint_status, notes, port_key,
                            service_history, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                        ON CONFLICT (id) DO NOTHING
                    """, (
                        a.get('id'), job_id,
                        a.get('tag',''), a.get('name',''), a.get('location',''),
                        a.get('eqType',''), a.get('manufacturer',''), a.get('model',''),
                        a.get('serial',''), a.get('voltage',''), a.get('criticality',''),
                        a.get('maintStatus',''), a.get('notes',''), a.get('portKey',''),
                        json.dumps([{
                            'yearId': year_id,
                            'label': 'C1-Y1',
                            'ir': a.get('ir',{}),
                            'ut': a.get('ut',{}),
                            'energized': a.get('energized',{}),
                            'insulation': a.get('insulation',{}),
                            'dlro': a.get('dlro',{}),
                            'torque': a.get('torque',{}),
                            'codeViolations': a.get('codeViolations',[]),
                        }]),
                    ))
            if old_jobs:
                print(f"Migrated {len(old_jobs)} old jobs to new structure")

        conn.commit(); cur.close(); conn.close()
        print("Database v3 initialized")
    except Exception as e:
        print(f"DB init error: {e}")
        import traceback; traceback.print_exc()

with app.app_context():
    init_db()

# ── Helpers ───────────────────────────────────────────────────
def sanitize_filename(name):
    safe = re.sub(r"[^a-zA-Z0-9 _\-]", "_", name or "Job")
    return safe.strip("_ ") or "Job"

def get_building_name(buildings, building_id):
    if not building_id: return "Main Bldg."
    for b in (buildings or []):
        if b.get("id") == building_id:
            return b.get("name", "Main Bldg.")
    return "Main Bldg."

EQUIPMENT_ORDER = [
    "swgr_main","swgr_icb","swgr_dob","swbd","swbd_disc","swbd_mcb",
    "panel_dist","panel_branch","mcc","mcc_starter","mcc_mcb","mcc_disc",
    "mcp","vfd_mcp","disc_ind","mcb_ind","starter_ind","xfmr_oil_mv",
    "xfmr_oil_hv","xfmr_dry","pf_cap","ltg_cont","pdu","wireway","busduct",
    "ats","mts","vfd_sm","vfd_lg","motor_sm","motor_md","motor_lg",
    "comp","hvac_lg","hvac_sm","ups_sm","ups_lg",
]
EQ_COUNT = len(EQUIPMENT_ORDER)

SVC_COLS = {
    "basic":    [4],
    "full":     [5],
    "full_af":  [5, 9],
    "full_den": [5, 6],
    "full_all": [5, 6, 9],
}

# ── Health ────────────────────────────────────────────────────
@app.route("/health", methods=["GET"])
def health():
    return jsonify({
        "status": "ok",
        "version": "3.0",
        "template": os.path.exists(TEMPLATE_PATH),
        "epm_template": os.path.exists(EPM_TEMPLATE_PATH),
        "database": DATABASE_URL is not None
    })

# ══════════════════════════════════════════════════════════════
# SITES
# ══════════════════════════════════════════════════════════════

@app.route("/sites", methods=["GET"])
def list_sites():
    if not DATABASE_URL: return jsonify([])
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            SELECT s.id, s.customer, s.address, s.device_name, s.updated_at,
                   COUNT(DISTINCT c.id) as contract_count,
                   COUNT(DISTINCT a.id) as asset_count
            FROM sites s
            LEFT JOIN contracts c ON c.site_id = s.id AND c.deleted = FALSE
            LEFT JOIN assets a ON a.site_id = s.id AND a.deleted = FALSE
            WHERE s.deleted = FALSE
            GROUP BY s.id, s.customer, s.address, s.device_name, s.updated_at
            ORDER BY s.updated_at DESC
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{
            "id": r[0], "customer": r[1], "address": r[2],
            "deviceName": r[3],
            "updatedAt": r[4].isoformat() if r[4] else None,
            "contractCount": r[5], "assetCount": r[6]
        } for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/sites/<site_id>", methods=["GET"])
def get_site(site_id):
    if not DATABASE_URL: return jsonify({"error": "No DB"}), 503
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT id, customer, address, contact, phone, notes, device_id, device_name, created_at, updated_at FROM sites WHERE id=%s AND deleted=FALSE", (site_id,))
        s = cur.fetchone()
        if not s: return jsonify({"error": "Not found"}), 404

        # Get contracts
        cur.execute("""
            SELECT id, contract_number, tier, status, start_date, end_date, notes, buildings, rooms, created_at, updated_at
            FROM contracts WHERE site_id=%s AND deleted=FALSE ORDER BY contract_number
        """, (site_id,))
        contracts = []
        for c in cur.fetchall():
            # Get years for this contract
            cur.execute("""
                SELECT id, year_number, label, work_date, rep, status, scoped_asset_ids, created_at, updated_at
                FROM years WHERE contract_id=%s AND deleted=FALSE ORDER BY year_number
            """, (c[0],))
            years = [{"id":y[0],"yearNumber":y[1],"label":y[2],"workDate":y[3],
                      "rep":y[4],"status":y[5],"scopedAssetIds":y[6] or [],
                      "createdAt":y[7].isoformat() if y[7] else None,
                      "updatedAt":y[8].isoformat() if y[8] else None} for y in cur.fetchall()]
            contracts.append({
                "id":c[0],"contractNumber":c[1],"tier":c[2],"status":c[3],
                "startDate":c[4],"endDate":c[5],"notes":c[6],
                "buildings":c[7] or [],"rooms":c[8] or [],
                "years":years,
                "createdAt":c[9].isoformat() if c[9] else None,
                "updatedAt":c[10].isoformat() if c[10] else None,
            })

        cur.close(); conn.close()
        return jsonify({
            "id":s[0],"customer":s[1],"address":s[2],"contact":s[3],"phone":s[4],
            "notes":s[5],"deviceId":s[6],"deviceName":s[7],
            "createdAt":s[8].isoformat() if s[8] else None,
            "updatedAt":s[9].isoformat() if s[9] else None,
            "contracts": contracts,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/sites/save", methods=["POST"])
def save_site():
    if not DATABASE_URL: return jsonify({"error": "No DB"}), 503
    try:
        payload = request.get_json(force=True)
        s = payload.get("site", {})
        device_id = payload.get("deviceId","")
        device_name = payload.get("deviceName","Unnamed Device")
        if not s.get("id"): return jsonify({"error": "No site id"}), 400

        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO sites (id, customer, address, contact, phone, notes, device_id, device_name, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            ON CONFLICT (id) DO UPDATE SET
                customer=EXCLUDED.customer, address=EXCLUDED.address,
                contact=EXCLUDED.contact, phone=EXCLUDED.phone,
                notes=EXCLUDED.notes, device_id=EXCLUDED.device_id,
                device_name=EXCLUDED.device_name, updated_at=NOW()
        """, (s["id"],s.get("customer",""),s.get("address",""),
              s.get("contact",""),s.get("phone",""),s.get("notes",""),
              device_id, device_name))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status":"saved","id":s["id"]})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/sites/<site_id>", methods=["DELETE"])
def delete_site(site_id):
    if not DATABASE_URL: return jsonify({"error": "No DB"}), 503
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("UPDATE sites SET deleted=TRUE, deleted_at=NOW() WHERE id=%s", (site_id,))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status":"deleted"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/sites/<site_id>/restore", methods=["POST"])
def restore_site(site_id):
    if not DATABASE_URL: return jsonify({"error": "No DB"}), 503
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("UPDATE sites SET deleted=FALSE, deleted_at=NULL WHERE id=%s", (site_id,))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status":"restored"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/sites/deleted", methods=["GET"])
def list_deleted_sites():
    if not DATABASE_URL: return jsonify([])
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT id, customer, address, device_name, deleted_at FROM sites WHERE deleted=TRUE ORDER BY deleted_at DESC")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{"id":r[0],"customer":r[1],"address":r[2],"deviceName":r[3],
                         "deletedAt":r[4].isoformat() if r[4] else None} for r in rows])
    except Exception as e:
        return jsonify({"error":str(e)}), 500

# ══════════════════════════════════════════════════════════════
# CONTRACTS
# ══════════════════════════════════════════════════════════════

@app.route("/contracts/save", methods=["POST"])
def save_contract():
    if not DATABASE_URL: return jsonify({"error": "No DB"}), 503
    try:
        c = request.get_json(force=True)
        if not c.get("id") or not c.get("siteId"): return jsonify({"error":"Missing fields"}), 400
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO contracts (id, site_id, contract_number, tier, status, start_date, end_date, notes, buildings, rooms, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            ON CONFLICT (id) DO UPDATE SET
                tier=EXCLUDED.tier, status=EXCLUDED.status,
                start_date=EXCLUDED.start_date, end_date=EXCLUDED.end_date,
                notes=EXCLUDED.notes, buildings=EXCLUDED.buildings,
                rooms=EXCLUDED.rooms, updated_at=NOW()
        """, (c["id"],c["siteId"],c.get("contractNumber",1),c.get("tier","basic"),
              c.get("status","Active"),c.get("startDate",""),c.get("endDate",""),
              c.get("notes",""),json.dumps(c.get("buildings",[])),json.dumps(c.get("rooms",[]))))
        # Update site updated_at
        cur.execute("UPDATE sites SET updated_at=NOW() WHERE id=%s", (c["siteId"],))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status":"saved","id":c["id"]})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/contracts/<contract_id>", methods=["DELETE"])
def delete_contract(contract_id):
    if not DATABASE_URL: return jsonify({"error": "No DB"}), 503
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("UPDATE contracts SET deleted=TRUE, deleted_at=NOW() WHERE id=%s", (contract_id,))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status":"deleted"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

# ══════════════════════════════════════════════════════════════
# YEARS
# ══════════════════════════════════════════════════════════════

@app.route("/years/save", methods=["POST"])
def save_year():
    if not DATABASE_URL: return jsonify({"error": "No DB"}), 503
    try:
        y = request.get_json(force=True)
        if not y.get("id"): return jsonify({"error":"No id"}), 400
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO years (id, contract_id, site_id, year_number, label, work_date, rep, status, scoped_asset_ids, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            ON CONFLICT (id) DO UPDATE SET
                work_date=EXCLUDED.work_date, rep=EXCLUDED.rep,
                status=EXCLUDED.status, scoped_asset_ids=EXCLUDED.scoped_asset_ids,
                updated_at=NOW()
        """, (y["id"],y.get("contractId"),y.get("siteId"),y.get("yearNumber",1),
              y.get("label","C1-Y1"),y.get("workDate",""),y.get("rep",""),
              y.get("status","Scheduled"),json.dumps(y.get("scopedAssetIds",[]))))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status":"saved","id":y["id"]})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

# ══════════════════════════════════════════════════════════════
# ASSETS
# ══════════════════════════════════════════════════════════════

@app.route("/sites/<site_id>/assets", methods=["GET"])
def get_assets(site_id):
    if not DATABASE_URL: return jsonify([])
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            SELECT id, tag, name, location, eq_type, manufacturer, model, serial,
                   voltage, criticality, maint_status, notes, port_key, service_history, updated_at
            FROM assets WHERE site_id=%s AND deleted=FALSE ORDER BY tag
        """, (site_id,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{
            "id":r[0],"tag":r[1],"name":r[2],"location":r[3],"eqType":r[4],
            "manufacturer":r[5],"model":r[6],"serial":r[7],"voltage":r[8],
            "criticality":r[9],"maintStatus":r[10],"notes":r[11],"portKey":r[12],
            "serviceHistory":r[13] or [],
            "updatedAt":r[14].isoformat() if r[14] else None
        } for r in rows])
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/assets/save", methods=["POST"])
def save_asset():
    if not DATABASE_URL: return jsonify({"error": "No DB"}), 503
    try:
        payload = request.get_json(force=True)
        assets = payload.get("assets", [])
        site_id = payload.get("siteId")
        if not site_id: return jsonify({"error":"No siteId"}), 400

        conn = get_db(); cur = conn.cursor()
        for a in assets:
            if not a.get("id"): continue
            cur.execute("""
                INSERT INTO assets (id, site_id, tag, name, location, eq_type, manufacturer,
                    model, serial, voltage, criticality, maint_status, notes, port_key,
                    service_history, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                ON CONFLICT (id) DO UPDATE SET
                    tag=EXCLUDED.tag, name=EXCLUDED.name, location=EXCLUDED.location,
                    eq_type=EXCLUDED.eq_type, manufacturer=EXCLUDED.manufacturer,
                    model=EXCLUDED.model, serial=EXCLUDED.serial, voltage=EXCLUDED.voltage,
                    criticality=EXCLUDED.criticality, maint_status=EXCLUDED.maint_status,
                    notes=EXCLUDED.notes, service_history=EXCLUDED.service_history,
                    updated_at=NOW()
            """, (a["id"],site_id,a.get("tag",""),a.get("name",""),a.get("location",""),
                  a.get("eqType",""),a.get("manufacturer",""),a.get("model",""),
                  a.get("serial",""),a.get("voltage",""),a.get("criticality",""),
                  a.get("maintStatus",""),a.get("notes",""),a.get("portKey",""),
                  json.dumps(a.get("serviceHistory",[]))))
        cur.execute("UPDATE sites SET updated_at=NOW() WHERE id=%s", (site_id,))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status":"saved","count":len(assets)})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/assets/<asset_id>", methods=["DELETE"])
def delete_asset(asset_id):
    if not DATABASE_URL: return jsonify({"error": "No DB"}), 503
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("UPDATE assets SET deleted=TRUE WHERE id=%s", (asset_id,))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status":"deleted"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

# ══════════════════════════════════════════════════════════════
# TAG COUNTER
# ══════════════════════════════════════════════════════════════

_tag_counter = {'value': None}

@app.route("/tag-counter", methods=["GET"])
def get_tag_counter():
    if not DATABASE_URL: return jsonify({"counter": _tag_counter['value']})
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT value FROM meta WHERE key='tag_counter'")
        row = cur.fetchone()
        cur.close(); conn.close()
        return jsonify({"counter": row[0] if row else None})
    except Exception as e:
        return jsonify({"counter": None})

@app.route("/tag-counter", methods=["POST"])
def set_tag_counter():
    data = request.get_json(force=True)
    counter = data.get('counter')
    if not counter: return jsonify({"status":"ignored"})
    _tag_counter['value'] = counter
    if not DATABASE_URL: return jsonify({"status":"ok"})
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO meta (key, value) VALUES ('tag_counter', %s)
            ON CONFLICT (key) DO UPDATE SET value = GREATEST(meta.value, EXCLUDED.value)
        """, (int(counter),))
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status":"ok"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

# ══════════════════════════════════════════════════════════════
# SALES BUILDER EXPORT
# ══════════════════════════════════════════════════════════════

@app.route("/export/sales", methods=["POST"])
def export_sales():
    try:
        payload = request.get_json(force=True)
        contract = payload.get("contract", {})
        site = payload.get("site", {})
        if not contract.get("rooms"): return jsonify({"error":"No rooms"}), 400
        if len(contract.get("rooms",[])) > 30: return jsonify({"error":"Max 30 rooms"}), 400

        wb = load_workbook(TEMPLATE_PATH)
        ws_br = wb["Bldgs & Rooms"]
        ws_c1 = wb["Components Yr 1"]

        buildings = contract.get("buildings", [])
        rooms = contract.get("rooms", [])
        for idx, room in enumerate(rooms):
            row_num = idx + 2
            bldg_name = get_building_name(buildings, room.get("buildingId"))
            ws_br.cell(row=row_num, column=1, value=bldg_name)
            ws_br.cell(row=row_num, column=2, value=room.get("name",""))

        default_svc = contract.get("tier","basic")
        for room_idx, room in enumerate(rooms):
            equipment = room.get("equipment", {})
            overrides = room.get("overrides", {})
            for eq_idx, eq_id in enumerate(EQUIPMENT_ORDER):
                qty = equipment.get(eq_id, 0)
                if not qty: continue
                svc = overrides.get(eq_id, default_svc)
                col_indices = SVC_COLS.get(svc, [4])
                row_num = (room_idx * EQ_COUNT) + eq_idx + 2
                for col_idx in col_indices:
                    ws_c1.cell(row=row_num, column=col_idx+1, value=qty)

        # Write pricing to Totals sheet yellow cells
        pricing = payload.get("pricing", {})
        if pricing and "Totals" in wb.sheetnames:
            ws_tot = wb["Totals"]
            try:
                faf_y1 = float(pricing.get("fafY1", 1.00))
                faf_y2 = float(pricing.get("fafY2", 1.00))
                faf_y3 = float(pricing.get("fafY3", 1.00))
                nfpa   = int(float(pricing.get("nfpaSeats", 0) or 0))
                margin = pricing.get("targetMargin", "")
                # Facility Access Factor — col D rows 6,7,8 (Y1,Y2,Y3)
                ws_tot.cell(row=6, column=4, value=faf_y1)
                ws_tot.cell(row=7, column=4, value=faf_y2)
                ws_tot.cell(row=8, column=4, value=faf_y3)
                # NFPA 70E seats — col N row 6 (Y1 only)
                ws_tot.cell(row=6, column=14, value=nfpa)
                # Target margin — col P rows 6,7,8
                if margin:
                    margin_val = float(margin) / 100
                    ws_tot.cell(row=6, column=16, value=margin_val)
                    ws_tot.cell(row=7, column=16, value=margin_val)
                    ws_tot.cell(row=8, column=16, value=margin_val)
            except Exception as pe:
                app.logger.warning(f"Pricing write error: {pe}")

        buffer = io.BytesIO()
        wb.save(buffer); buffer.seek(0)
        customer = site.get("customer","Job")
        cn = contract.get("contractNumber",1)
        filename = sanitize_filename(customer) + f"_C{cn}_EMA_Sales_Builder.xlsx"
        return send_file(buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=filename)
    except Exception as e:
        app.logger.error(f"Sales export error: {e}", exc_info=True)
        return jsonify({"error":str(e)}), 500

# ══════════════════════════════════════════════════════════════
# EPM MASTER EXPORT
# ══════════════════════════════════════════════════════════════

def na_or(val, default="N/A"):
    """Return value or default if empty."""
    return val if val else default

@app.route("/export/epm", methods=["POST"])
def export_epm():
    try:
        payload = request.get_json(force=True)
        site = payload.get("site", {})
        year = payload.get("year", {})
        assets = payload.get("assets", [])
        year_id = year.get("id","")
        label = year.get("label","Y1")

        if not assets: return jsonify({"error":"No assets"}), 400

        # Get field data for this specific year from service history
        def get_year_data(asset, field):
            for entry in (asset.get("serviceHistory") or []):
                if entry.get("yearId") == year_id:
                    return entry.get(field, {})
            # Fall back to top-level field data (legacy)
            return asset.get(field, {})

        def get_year_violations(asset):
            for entry in (asset.get("serviceHistory") or []):
                if entry.get("yearId") == year_id:
                    return entry.get("codeViolations", [])
            return asset.get("codeViolations", [])

        template = EPM_TEMPLATE_PATH if os.path.exists(EPM_TEMPLATE_PATH) else TEMPLATE_PATH
        wb = load_workbook(template)

        # Equipment Information
        if "Equipment Information" in wb.sheetnames:
            ws = wb["Equipment Information"]
            for idx, a in enumerate(assets, start=2):
                ws.cell(row=idx, column=1,  value=a.get("tag",""))
                ws.cell(row=idx, column=2,  value=a.get("name",""))
                ws.cell(row=idx, column=3,  value=a.get("location",""))
                ws.cell(row=idx, column=4,  value=a.get("eqType",""))
                ws.cell(row=idx, column=5,  value=a.get("manufacturer",""))
                ws.cell(row=idx, column=6,  value=a.get("model",""))
                ws.cell(row=idx, column=7,  value=a.get("serial",""))
                ws.cell(row=idx, column=8,  value=a.get("voltage",""))
                ws.cell(row=idx, column=9,  value=a.get("criticality",""))
                ws.cell(row=idx, column=10, value=a.get("maintStatus",""))
                ws.cell(row=idx, column=13, value=a.get("notes",""))

        # Code Violations
        if "Code Violations" in wb.sheetnames:
            ws = wb["Code Violations"]
            for idx, a in enumerate(assets, start=2):
                ws.cell(row=idx, column=1, value=a.get("tag",""))
                ws.cell(row=idx, column=2, value=a.get("name",""))
                ws.cell(row=idx, column=3, value=a.get("location",""))
                violations = get_year_violations(a)
                for vi, v in enumerate(violations[:5]):
                    if not v.get("article"): continue
                    col_base = 4 + (vi * 3)
                    ws.cell(row=idx, column=col_base,   value=v.get("article",""))
                    ws.cell(row=idx, column=col_base+1, value=v.get("sev",""))
                    ws.cell(row=idx, column=col_base+2, value=v.get("corrected","No"))

        # Infrared
        if "Infrared" in wb.sheetnames:
            ws = wb["Infrared"]
            for idx, a in enumerate(assets, start=2):
                ir = get_year_data(a, "ir")
                ws.cell(row=idx, column=1, value=a.get("tag",""))
                ws.cell(row=idx, column=2, value=a.get("name",""))
                ws.cell(row=idx, column=3, value=a.get("location",""))
                ws.cell(row=idx, column=4, value=ir.get("phase",""))
                ws.cell(row=idx, column=5, value=ir.get("component",""))
                ws.cell(row=idx, column=6, value=na_or(ir.get("result","")))
                ws.cell(row=idx, column=7, value=ir.get("notes",""))

        # Ultrasonic
        if "Ultrasonic" in wb.sheetnames:
            ws = wb["Ultrasonic"]
            for idx, a in enumerate(assets, start=2):
                ut = get_year_data(a, "ut")
                ws.cell(row=idx, column=1, value=a.get("tag",""))
                ws.cell(row=idx, column=2, value=a.get("name",""))
                ws.cell(row=idx, column=3, value=a.get("location",""))
                ws.cell(row=idx, column=4, value=ut.get("component",""))
                ws.cell(row=idx, column=5, value=ut.get("failureMode",""))
                ws.cell(row=idx, column=6, value=ut.get("db",""))
                ws.cell(row=idx, column=7, value=na_or(ut.get("result","")))
                ws.cell(row=idx, column=8, value=ut.get("notes",""))

        # Energized Data
        if "Energized Data" in wb.sheetnames:
            ws = wb["Energized Data"]
            for idx, a in enumerate(assets, start=2):
                en = get_year_data(a, "energized")
                ws.cell(row=idx, column=1,  value=a.get("tag",""))
                ws.cell(row=idx, column=2,  value=a.get("name",""))
                ws.cell(row=idx, column=3,  value=a.get("location",""))
                ws.cell(row=idx, column=4,  value=en.get("phase",""))
                ws.cell(row=idx, column=5,  value=en.get("wire",""))
                ws.cell(row=idx, column=6,  value=en.get("vab",""))
                ws.cell(row=idx, column=7,  value=en.get("vbc",""))
                ws.cell(row=idx, column=8,  value=en.get("vca",""))
                ws.cell(row=idx, column=10, value=en.get("aa",""))
                ws.cell(row=idx, column=11, value=en.get("ab",""))
                ws.cell(row=idx, column=12, value=en.get("ac",""))
                ws.cell(row=idx, column=14, value=en.get("ratedAmps",""))
                ws.cell(row=idx, column=15, value=en.get("ocpd",""))
                ws.cell(row=idx, column=18, value=na_or(en.get("analysis","")))
                ws.cell(row=idx, column=19, value=na_or(en.get("summary","")))

        # Insulation
        if "Insulation" in wb.sheetnames:
            ws = wb["Insulation"]
            for idx, a in enumerate(assets, start=2):
                ins = get_year_data(a, "insulation")
                ws.cell(row=idx, column=1, value=a.get("tag",""))
                ws.cell(row=idx, column=2, value=a.get("name",""))
                ws.cell(row=idx, column=3, value=a.get("location",""))
                ws.cell(row=idx, column=4, value=ins.get("testV",""))
                ws.cell(row=idx, column=5, value=ins.get("phPh",""))
                ws.cell(row=idx, column=6, value=ins.get("phG",""))
                ws.cell(row=idx, column=7, value=ins.get("pi",""))
                ws.cell(row=idx, column=8, value=ins.get("temp",""))
                ws.cell(row=idx, column=9, value=na_or(ins.get("result","")))
                ws.cell(row=idx, column=10, value=ins.get("notes",""))

        # DLRO
        if "DLRO" in wb.sheetnames:
            ws = wb["DLRO"]
            for idx, a in enumerate(assets, start=2):
                dl = get_year_data(a, "dlro")
                ws.cell(row=idx, column=1, value=a.get("tag",""))
                ws.cell(row=idx, column=2, value=a.get("name",""))
                ws.cell(row=idx, column=3, value=a.get("location",""))
                ws.cell(row=idx, column=4, value=dl.get("phA",""))
                ws.cell(row=idx, column=5, value=dl.get("phB",""))
                ws.cell(row=idx, column=6, value=dl.get("phC",""))
                ws.cell(row=idx, column=7, value=dl.get("dev",""))
                ws.cell(row=idx, column=8, value=na_or(dl.get("result","")))
                ws.cell(row=idx, column=9, value=dl.get("notes",""))

        # Torque
        if "Torque" in wb.sheetnames:
            ws = wb["Torque"]
            for idx, a in enumerate(assets, start=2):
                tq = get_year_data(a, "torque")
                ws.cell(row=idx, column=1, value=a.get("tag",""))
                ws.cell(row=idx, column=2, value=a.get("name",""))
                ws.cell(row=idx, column=3, value=a.get("location",""))
                ws.cell(row=idx, column=4, value=tq.get("phase",""))
                ws.cell(row=idx, column=5, value=tq.get("connType",""))
                ws.cell(row=idx, column=6, value=tq.get("spec",""))
                ws.cell(row=idx, column=7, value=tq.get("verified",""))
                ws.cell(row=idx, column=8, value=na_or(tq.get("result","")))
                ws.cell(row=idx, column=9, value=tq.get("notes",""))

        buffer = io.BytesIO()
        wb.save(buffer); buffer.seek(0)
        customer = site.get("customer","Job")
        filename = sanitize_filename(customer) + f"_{label}_EPM_Master.xlsx"
        return send_file(buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=filename)
    except Exception as e:
        app.logger.error(f"EPM export error: {e}", exc_info=True)
        return jsonify({"error":str(e)}), 500

# ── Backward compat — old /export and /jobs routes ────────────
@app.route("/export", methods=["POST"])
def export_legacy():
    """Legacy route — redirect to new sales export."""
    try:
        job = request.get_json(force=True)
        return export_sales.__wrapped__({"contract": job, "site": job}) if hasattr(export_sales, '__wrapped__') else export_sales()
    except:
        return jsonify({"error":"Use /export/sales"}), 400

@app.route("/jobs", methods=["GET"])
def list_jobs_legacy():
    """Legacy — return sites list in old format."""
    return list_sites()

@app.route("/jobs/save", methods=["POST"])
def save_job_legacy():
    """Legacy — save as site."""
    try:
        payload = request.get_json(force=True)
        job = payload.get("job", payload)
        new_payload = {
            "site": {
                "id": job.get("id"),
                "customer": job.get("customer",""),
                "address": job.get("address",""),
                "notes": job.get("notes",""),
            },
            "deviceId": payload.get("deviceId", job.get("deviceId","")),
            "deviceName": payload.get("deviceName", job.get("deviceName",""))
        }
        with app.test_request_context(json=new_payload):
            request._cached_json = (new_payload, new_payload)
        return save_site()
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/jobs/<job_id>", methods=["GET"])
def get_job_legacy(job_id):
    return get_site(job_id)


# ── One-time cleanup ─────────────────────────────────────────
@app.route("/admin/clear-migrated", methods=["POST"])
def clear_migrated():
    """Delete all records migrated from old job format (id starts with job_)"""
    if not DATABASE_URL:
        return jsonify({"error": "No DB"}), 503
    try:
        conn = get_db(); cur = conn.cursor()
        # Delete assets from migrated sites
        cur.execute("DELETE FROM assets WHERE site_id LIKE 'job_%'")
        assets_deleted = cur.rowcount
        # Delete years from migrated contracts
        cur.execute("DELETE FROM years WHERE contract_id LIKE 'c1_job_%'")
        years_deleted = cur.rowcount
        # Delete migrated contracts
        cur.execute("DELETE FROM contracts WHERE id LIKE 'c1_job_%'")
        contracts_deleted = cur.rowcount
        # Delete migrated sites
        cur.execute("DELETE FROM sites WHERE id LIKE 'job_%'")
        sites_deleted = cur.rowcount
        conn.commit(); cur.close(); conn.close()
        return jsonify({
            "status": "cleared",
            "sites": sites_deleted,
            "contracts": contracts_deleted,
            "years": years_deleted,
            "assets": assets_deleted
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/admin/wipe-all", methods=["POST"])
def wipe_all():
    """Wipe all data for a clean start"""
    if not DATABASE_URL:
        return jsonify({"error": "No DB"}), 503
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("DELETE FROM assets")
        cur.execute("DELETE FROM years")
        cur.execute("DELETE FROM contracts")
        cur.execute("DELETE FROM sites")
        cur.execute("DELETE FROM meta")
        conn.commit(); cur.close(); conn.close()
        return jsonify({"status": "wiped", "message": "All data cleared"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
